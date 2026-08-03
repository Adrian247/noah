#!/usr/bin/env python3
"""Convierte bitácoras de mantenimiento en Excel al contrato de ingesta de Phoenix.

Phoenix no lee `.xlsm`: el ETL vive aquí y produce un JSON estable que consume
`php artisan phoenix:predictive:ingest`. Así se pueden agregar nuevos formatos de bitácora
sin tocar el monolito.

Perfiles soportados:

* ``plant``       — bitácora de planta de proceso (hojas Data, Equipment, Events, Alarms, OT,
                    Milestone). Un renglón por equipo y día.
* ``underground`` — reporte diario de mina (hojas Datos, Equipos). Un renglón por equipo y turno.

Uso::

    python scripts/extract_logbooks.py "Bitácora Planta 4400.xlsm" \
        --profile plant --site "Planta 4400" -o datasets/planta-4400.json
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook

# El libro codifica la clave del renglón como <tag><serialExcel><turno>, p. ej. JB-10044075T2.
ROW_KEY_RE = re.compile(r"^(?P<tag>.+?)(?P<serial>4\d{4})(?P<shift>T\d)$")

EXCEL_EPOCH = dt.date(1899, 12, 30)

UNDERGROUND_CLASSES = {
    "JB": ("JUMBO", "Jumbo de barrenación frontal"),
    "SS": ("SCOOPTRAM", "Cargador frontal de bajo perfil (LHD)"),
    "VQ": ("CAMION_BAJO_PERFIL", "Camión de bajo perfil"),
}

# Se clasifica por palabra clave del nombre porque la bitácora de planta no trae clase.
PLANT_CLASS_KEYWORDS = [
    ("QUEBRADORA", "QUEBRADORA"),
    ("TRITURADORA", "QUEBRADORA"),
    ("MOLINO", "MOLINO"),
    ("ESPESADOR", "ESPESADOR"),
    ("FILTRO", "FILTRO"),
    ("CRIBA", "CRIBA"),
    ("CELDAS", "CELDA_FLOTACION"),
    ("BANCO", "CELDA_FLOTACION"),
    ("ACONDICIONADOR", "ACONDICIONADOR"),
    ("ALIMENTADOR", "ALIMENTADOR"),
    ("BANDA", "BANDA_TRANSPORTADORA"),
    ("TRANSPORTADOR", "BANDA_TRANSPORTADORA"),
    ("ELEVADOR", "ELEVADOR"),
    ("BOMBA", "BOMBA"),
    ("COMPRESOR", "COMPRESOR"),
    ("VENTILADOR", "VENTILADOR"),
    ("SOPLADOR", "VENTILADOR"),
    ("COLECTOR", "COLECTOR_POLVOS"),
    ("MOTOR", "MOTOR_ELECTRICO"),
    ("TANQUE", "TANQUE"),
    ("GRUA", "GRUA"),
    ("SECADOR", "SECADOR"),
]

MANUFACTURER_KEYWORDS = [
    ("METSO", "Metso"),
    ("NORDBERG", "Metso"),
    ("OUTOKUMPU", "Outotec"),
    ("OUTOTEC", "Outotec"),
    ("EIMCO", "Eimco"),
    ("ALLIS CHALMER", "Allis-Chalmers"),
    ("DENVER", "Denver"),
    ("WEMCO", "Wemco"),
    ("TYCAN", "Tycan"),
    ("CAMFIL", "Camfil"),
    ("EPIROC", "Epiroc"),
    ("ATLAS COPCO", "Epiroc"),
    ("SANDVIK", "Sandvik"),
]

MODEL_RE = re.compile(r"\b(?:HP|C|MP|GP|LT|DD|DL|DS|LH|TH|ST|MT)-?\d{2,4}[A-Za-z]?\b")


# --------------------------------------------------------------------------------------
# Utilidades de lectura
# --------------------------------------------------------------------------------------


def read_sheet(path: Path, sheet: str, header_row: int = 1) -> list[dict[str, Any]]:
    """Devuelve los renglones de una hoja como dicts, ignorando renglones totalmente vacíos."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet]
        header: list[str] = []
        rows: list[dict[str, Any]] = []
        for index, raw in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index < header_row:
                continue
            if index == header_row:
                header = [str(cell).strip() if cell is not None else "" for cell in raw]
                continue
            if all(cell is None for cell in raw):
                continue
            rows.append(dict(zip(header, raw)))
        return rows
    finally:
        workbook.close()


def text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = re.sub(r"\s+", " ", str(value)).strip()
    return cleaned or None


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    try:
        return round(float(str(value).replace(",", "").strip()), 4)
    except (TypeError, ValueError):
        return None


def as_date(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 30000 < float(value) < 60000:
        return (EXCEL_EPOCH + dt.timedelta(days=int(value))).isoformat()
    return None


def as_time(value: Any) -> dt.time | None:
    if isinstance(value, dt.time):
        return value
    if isinstance(value, dt.datetime):
        return value.time()
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        seconds = int(round(float(value) * 86400))
        return (dt.datetime.min + dt.timedelta(seconds=seconds)).time()
    return None


def combine(day: str | None, moment: dt.time | None) -> str | None:
    if day is None:
        return None
    base = dt.date.fromisoformat(day)
    return dt.datetime.combine(base, moment or dt.time()).isoformat(sep=" ")


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", ascii_only.lower()).strip("-")


def parse_row_key(value: Any) -> tuple[str, str, str] | None:
    """Extrae ``(tag, fecha ISO, turno)`` de la clave compuesta del libro."""
    raw = text(value)
    if raw is None:
        return None
    match = ROW_KEY_RE.match(raw.replace(" ", ""))
    if match is None:
        return None
    day = (EXCEL_EPOCH + dt.timedelta(days=int(match.group("serial")))).isoformat()
    return match.group("tag"), day, match.group("shift")


def positive(value: Any) -> float:
    parsed = number(value)
    return parsed if parsed and parsed > 0 else 0.0


def classify_plant(name: str | None) -> str:
    upper = (name or "").upper()
    for keyword, equipment_class in PLANT_CLASS_KEYWORDS:
        if keyword in upper:
            return equipment_class
    return "OTRO"


def detect_manufacturer(name: str | None) -> str | None:
    upper = (name or "").upper()
    for keyword, manufacturer in MANUFACTURER_KEYWORDS:
        if keyword in upper:
            return manufacturer
    return None


def detect_model(name: str | None) -> str | None:
    match = MODEL_RE.search((name or "").upper())
    return match.group(0) if match else None


def compact(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if value not in (None, "", [], {})}


def drop_date_outliers(
    rows: list[dict[str, Any]], key: str, tolerance_days: int = 45
) -> tuple[list[dict[str, Any]], list[str]]:
    """Descarta renglones cuya fecha está muy lejos del cuerpo del libro.

    Los libros traen fechas mal capturadas de forma aislada (dos renglones de septiembre 2020
    tecleados como enero 2020). Si se dejan, extienden el periodo de cobertura del equipo y
    envenenan cualquier cálculo de ventana o de etiqueta. Se compara contra la mediana porque es
    inmune a esos mismos valores atípicos.
    """
    days = sorted(row[key][:10] for row in rows if row.get(key))
    if len(days) < 10:
        return rows, []

    median = dt.date.fromisoformat(days[len(days) // 2])
    kept: list[dict[str, Any]] = []
    dropped: list[str] = []
    for row in rows:
        value = (row.get(key) or "")[:10]
        if value and abs((dt.date.fromisoformat(value) - median).days) > tolerance_days:
            dropped.append(value)
            continue
        kept.append(row)

    return kept, dropped


def merge_failure_episodes(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Une renglones de falla contiguos en un solo episodio.

    Las bitácoras registran horas de reparación turno por turno: una avería que dura tres turnos
    aparece como tres renglones. Contarlos como tres fallas infla la tasa y arruina el MTBF, así
    que los renglones del mismo equipo separados por un día o menos se consolidan en un episodio
    con el tiempo fuera de servicio sumado.
    """
    by_asset: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_asset.setdefault(row["asset_tag"], []).append(row)

    episodes: list[dict[str, Any]] = []
    for asset_rows in by_asset.values():
        asset_rows.sort(key=lambda r: r["started_at"])
        current: dict[str, Any] | None = None
        for row in asset_rows:
            started = dt.datetime.fromisoformat(row["started_at"])
            if current is not None:
                previous_day = dt.datetime.fromisoformat(current["_last_day"]).date()
                if (started.date() - previous_day).days <= 1:
                    current["downtime_hours"] = round(
                        (current.get("downtime_hours") or 0) + (row.get("downtime_hours") or 0), 2
                    )
                    current["ended_at"] = row.get("ended_at") or current.get("ended_at")
                    current["_last_day"] = row["started_at"]
                    current["shifts"] = current.get("shifts", 1) + 1
                    # Se conserva la descripción más informativa del episodio.
                    if len(row.get("reported_text") or "") > len(current.get("reported_text") or ""):
                        current["reported_text"] = row.get("reported_text")
                    continue
                episodes.append(current)
            current = {**row, "_last_day": row["started_at"], "shifts": 1}
        if current is not None:
            episodes.append(current)

    for episode in episodes:
        episode.pop("_last_day", None)

    episodes.sort(key=lambda r: (r["asset_tag"], r["started_at"]))
    return episodes


def dedupe(rows: list[dict[str, Any]], *keys: str) -> list[dict[str, Any]]:
    """Los libros originales repiten renglones idénticos; se conserva la primera aparición."""
    seen: set[tuple[Any, ...]] = set()
    unique: list[dict[str, Any]] = []
    for row in rows:
        signature = tuple(row.get(key) for key in keys)
        if signature in seen:
            continue
        seen.add(signature)
        unique.append(row)
    return unique


# --------------------------------------------------------------------------------------
# Perfil: planta de proceso
# --------------------------------------------------------------------------------------


def extract_plant(path: Path, site: str) -> dict[str, list[dict[str, Any]]]:
    assets: dict[str, dict[str, Any]] = {}

    def register(tag: str | None, name: str | None, **extra: Any) -> str | None:
        tag = text(tag)
        if tag is None:
            return None
        entry = assets.setdefault(
            tag,
            {
                "tag": tag,
                "name": name or tag,
                "equipment_class": classify_plant(name),
                "manufacturer": detect_manufacturer(name),
                "model": detect_model(name),
                "application": "surface",
            },
        )
        if name and (entry["name"] == tag or len(name) > len(entry["name"])):
            entry["name"] = name
            entry["equipment_class"] = classify_plant(name)
            entry["manufacturer"] = entry["manufacturer"] or detect_manufacturer(name)
            entry["model"] = entry["model"] or detect_model(name)
        entry.update({k: v for k, v in extra.items() if v is not None})
        return tag

    for row in read_sheet(path, "Equipment"):
        register(
            row.get("Tag"),
            text(row.get("Equipment")),
            area=text(row.get("Area")),
            location_code=text(row.get("Location")),
            hour_meter=number(row.get("Hours")),
            counter=number(row.get("Counter")),
            is_main=bool(row.get("isMain")),
        )

    shift_logs: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []

    for row in read_sheet(path, "Data"):
        key = parse_row_key(row.get("DNI"))
        tag = register(row.get("Tag"), text(row.get("Equipment")))
        if tag is None:
            continue
        day = as_date(row.get("Date")) or (key[1] if key else None)
        if day is None:
            continue
        shift = key[2] if key else "FULL"
        preventive = positive(row.get("M.Preventive"))
        corrective = positive(row.get("M.Corrective"))
        operative = positive(row.get("OperativeFail"))
        tons = number(row.get("TonsTrituration"))

        shift_logs.append(
            compact(
                {
                    "asset_tag": tag,
                    "logged_on": day,
                    "shift": shift,
                    "scheduled_hours": positive(row.get("ScheduledHours")),
                    "worked_hours": positive(row.get("HrWork")),
                    "standby_hours": positive(row.get("StandBy")),
                    "preventive_hours": preventive,
                    "corrective_hours": corrective,
                    "operative_fail_hours": operative,
                    "availability": number(row.get("Availability")),
                    "utilization": number(row.get("Utilization")),
                    "production": {"tons": tons} if tons else None,
                    "location_label": text(row.get("Area")),
                    "failure_text": text(row.get("Comments")),
                    "comments": text(row.get("CheckListComments")),
                    "source": "logbook_plant",
                    "external_ref": text(row.get("DNI")),
                }
            )
        )

        downtime = round(corrective + operative, 2)
        if downtime > 0:
            failures.append(
                compact(
                    {
                        "asset_tag": tag,
                        "started_at": f"{day} 00:00:00",
                        "downtime_hours": downtime,
                        "maintenance_type": "corrective" if corrective > 0 else "operational",
                        "reported_text": text(row.get("Comments")),
                        "source": "logbook_plant",
                        "external_ref": text(row.get("DNI")),
                    }
                )
            )

    events: list[dict[str, Any]] = []
    for sheet, count_column, source in (("Events", None, "plc"), ("Alarms", "AlarmCount", "plc_alarm")):
        for row in read_sheet(path, sheet):
            code = text(row.get("EventId") or row.get("AlarmId"))
            tag = register(row.get("Tag"), text(row.get("Equipment")))
            occurred = combine(as_date(row.get("Date")), as_time(row.get("Time")))
            if code is None or tag is None or occurred is None:
                continue
            events.append(
                compact(
                    {
                        "asset_tag": tag,
                        "occurred_at": occurred,
                        "code": code,
                        "name": text(row.get("EventName") or row.get("AlarmName")) or code,
                        "occurrences": int(number(row.get(count_column)) or 1) if count_column else 1,
                        "source": source,
                    }
                )
            )

    work_orders: list[dict[str, Any]] = []
    for row in read_sheet(path, "OT"):
        order_number = text(row.get("No. Orden"))
        if order_number is None:
            continue
        executed = (text(row.get("Ejecución")) or "").upper().startswith("S")
        planned = as_date(row.get("Fecha Planeada"))
        work_orders.append(
            compact(
                {
                    "asset_tag": text(row.get("No. Equipo")),
                    "order_number": order_number,
                    "description": text(row.get("Descripción del Servicio")),
                    "work_center": text(row.get("Denominación")),
                    "location_code": text(row.get("Ubicación")),
                    "planned_for": planned,
                    "executed_on": planned if executed else None,
                    "status": "executed" if executed else "skipped",
                    "skip_reason": None
                    if executed
                    else text(row.get("Motivo por lo que no se realizó mantenimiento")),
                    "supervisor": text(row.get("Supervisor")),
                    "source": "sap",
                }
            )
        )

    replacements: list[dict[str, Any]] = []
    for row in read_sheet(path, "Milestone"):
        tag = text(row.get("Tag"))
        replaced = combine(as_date(row.get("Date")), as_time(row.get("Time")))
        activity = text(row.get("Activity"))
        if tag is None or replaced is None or activity is None:
            continue
        replacements.append(
            compact(
                {
                    "asset_tag": tag,
                    "component": activity[:120],
                    "description": activity,
                    "replaced_at": replaced,
                    "source": "logbook_plant",
                }
            )
        )

    return {
        "site": site,
        "assets": list(assets.values()),
        "shift_logs": dedupe(shift_logs, "asset_tag", "logged_on", "shift"),
        "events": dedupe(events, "asset_tag", "occurred_at", "code", "source"),
        "failures": merge_failure_episodes(dedupe(failures, "asset_tag", "started_at", "maintenance_type")),
        "work_orders": dedupe(work_orders, "order_number"),
        "component_replacements": dedupe(replacements, "asset_tag", "component", "replaced_at"),
        "measurements": [],
    }


# --------------------------------------------------------------------------------------
# Perfil: mina subterránea
# --------------------------------------------------------------------------------------


def extract_underground(path: Path, site: str) -> dict[str, list[dict[str, Any]]]:
    assets: dict[str, dict[str, Any]] = {}

    for row in read_sheet(path, "Equipos", header_row=2):
        tag = text(row.get("NombreCorto"))
        class_code = (text(row.get("Clase")) or "").upper()
        if tag is None or class_code not in UNDERGROUND_CLASSES:
            continue
        equipment_class, description = UNDERGROUND_CLASSES[class_code]
        assets[tag] = compact(
            {
                "tag": tag,
                "name": f"{text(row.get('Descripcion')) or description} {tag}",
                "equipment_class": equipment_class,
                "application": "underground",
                "serial_number": text(row.get("Equipo")),
                "source_class": class_code,
                "capacity": number(row.get("Capacidad")),
            }
        )

    shift_logs: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []

    for row in read_sheet(path, "Datos", header_row=3):
        key = parse_row_key(row.get("Auxiliar"))
        tag = text(row.get("No Eco")) or (key[0] if key else None)
        if tag is None:
            continue
        # La fecha de la clave compuesta es la autoritativa: la columna Fecha trae renglones
        # desalineados en el libro original.
        day = (key[1] if key else None) or as_date(row.get("Fecha"))
        if day is None:
            continue
        shift = text(row.get("Turno")) or (key[2] if key else "T1")

        if tag not in assets:
            class_code = (text(row.get("Clase Equipo")) or "").upper()
            equipment_class, description = UNDERGROUND_CLASSES.get(class_code, ("OTRO", "Equipo"))
            assets[tag] = {
                "tag": tag,
                "name": f"{description} {tag}",
                "equipment_class": equipment_class,
                "application": "underground",
                "source_class": class_code or None,
            }

        corrective = positive(row.get("Falla / Reparacion"))
        production = compact(
            {
                "mineral_tons": number(row.get("Mineral")),
                "waste_tons": number(row.get("Tepetate")),
                "meters": number(row.get("Mts")),
                "holes": number(row.get("No Barrenos")),
                "cut": number(row.get("Corte")),
                "stripping": number(row.get("Desb.")),
                "bench": number(row.get("Banco")),
                "advance": number(row.get("Avance")),
                "bolts": number(row.get("Ancl.")),
            }
        )

        shift_logs.append(
            compact(
                {
                    "asset_tag": tag,
                    "logged_on": day,
                    "shift": shift,
                    "scheduled_hours": positive(row.get("HrsTotal")),
                    "worked_hours": positive(row.get("Horas Trab.")),
                    "standby_hours": positive(row.get("Tiempo Traslado")),
                    "preventive_hours": positive(row.get("Servicio")),
                    "corrective_hours": corrective,
                    "operative_fail_hours": positive(row.get("Por Oper.")),
                    "no_operator_hours": positive(row.get("Falta Operador")),
                    "availability": number(row.get("Disp.")),
                    "utilization": number(row.get("Util.")),
                    "hour_meter_start": number(row.get("H Incial")),
                    "hour_meter_end": number(row.get("H Final")),
                    "diesel_liters": number(row.get("Diesel")),
                    "oil_liters": number(row.get("Aceite")),
                    "coolant_liters": number(row.get("Refirgerante")),
                    "production": production or None,
                    "location_label": text(row.get("Lugar")),
                    "equipment_status": text(row.get("Estatus Equipo")),
                    "failure_text": text(row.get("Estatus / Falla")),
                    "source": "logbook_underground",
                    "external_ref": text(row.get("Auxiliar")),
                }
            )
        )

        if corrective > 0:
            started = combine(day, as_time(row.get("Paro Mantto. Inicio")))
            ended = combine(day, as_time(row.get("Paro Mantto. Fin")))
            failures.append(
                compact(
                    {
                        "asset_tag": tag,
                        "started_at": started or f"{day} 00:00:00",
                        "ended_at": ended,
                        "downtime_hours": corrective,
                        "maintenance_type": "corrective",
                        "reported_text": text(row.get("Estatus / Falla")),
                        "hour_meter": number(row.get("H Incial")),
                        "source": "logbook_underground",
                        "external_ref": text(row.get("Auxiliar")),
                    }
                )
            )

    return {
        "site": site,
        "assets": list(assets.values()),
        "shift_logs": dedupe(shift_logs, "asset_tag", "logged_on", "shift"),
        "events": [],
        "failures": merge_failure_episodes(dedupe(failures, "asset_tag", "started_at", "maintenance_type")),
        "work_orders": [],
        "component_replacements": [],
        "measurements": [],
    }


# --------------------------------------------------------------------------------------


EXTRACTORS = {"plant": extract_plant, "underground": extract_underground}


def build_dataset(path: Path, profile: str, site: str) -> dict[str, Any]:
    payload = EXTRACTORS[profile](path, site)

    payload["shift_logs"], stray_logs = drop_date_outliers(payload["shift_logs"], "logged_on")
    payload["failures"], stray_failures = drop_date_outliers(payload["failures"], "started_at")

    counts = {key: len(value) for key, value in payload.items() if isinstance(value, list)}
    return {
        "meta": {
            "contract": "phoenix.predictive.ingest/v1",
            "profile": profile,
            "source_file": path.name,
            "dataset_key": slug(f"{profile}-{path.stem}"),
            "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
            "counts": counts,
            "dropped_date_outliers": {
                "shift_logs": sorted(set(stray_logs)),
                "failures": sorted(set(stray_failures)),
            },
        },
        **payload,
    }


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("source", type=Path, help="Archivo .xlsm/.xlsx de la bitácora")
    parser.add_argument("--profile", choices=sorted(EXTRACTORS), required=True)
    parser.add_argument("--site", default="Sitio principal", help="Nombre del sitio destino")
    parser.add_argument("-o", "--output", type=Path, help="Ruta del JSON de salida")
    args = parser.parse_args(list(argv) if argv is not None else None)

    dataset = build_dataset(args.source, args.profile, args.site)
    output = args.output or Path("datasets") / f"{dataset['meta']['dataset_key']}.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(dataset, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"{output}")
    for key, total in dataset["meta"]["counts"].items():
        print(f"  {key}: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
